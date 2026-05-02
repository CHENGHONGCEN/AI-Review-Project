import base64
import io
import json
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from html import escape, unescape
from time import sleep
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI


DEFAULT_BASE_URL = "https://api.openai.com/v1"
DEFAULT_MODEL = "gpt-5.5"
CONFIDENCE_LEVELS = ["high", "medium", "low"]
SIMILARITY_THRESHOLD = 0.95
CITATION_ABSTRACT_TOKEN_OVERLAP_THRESHOLD = 0.95
CITATION_AI_MAX_BATCH_RECORDS = 8
CITATION_AI_BATCH_CHAR_BUDGET = 30000
CITATION_AI_BATCH_DELAY_SECONDS = 3
MMAT_RESPONSES = ["Yes", "No", "Can't tell"]
MMAT_STUDY_DESIGNS = [
    "Qualitative",
    "Quantitative randomized controlled trial",
    "Quantitative non-randomized",
    "Quantitative descriptive",
    "Mixed methods",
    "Not suitable for MMAT",
]
DEFAULT_PROMPT_TEMPLATE = """
You are an expert systematic review data extractor with experience in evidence synthesis, qualitative evidence extraction, health communication research, and thematic analysis.

Your task is to read the attached research article PDF and extract information for one article record. Work as a careful research assistant, not as a creative writer.

Core principles:
- Accuracy is more important than fluency.
- Transparency is more important than completeness of prose.
- Do not invent, infer beyond the article, or fill gaps from general knowledge.
- If the article does not clearly provide an answer, write "Not found".
- If the article is unrelated to a research question, write "Not relevant to this article" in the summary and explain why briefly when possible.
- Preserve original wording in evidence excerpts. Do not paraphrase inside excerpt text.
- Keep summaries concise and evidence-based. Summaries must be supported by the extracted excerpts or clearly identifiable article content.

Anti-hallucination rules:
- Use only information found in the attached PDF.
- Do not use outside knowledge, assumptions about the topic, or common patterns from similar papers.
- Do not make up page numbers, sections, participant characteristics, methods, outcomes, or conclusions.
- Do not treat the abstract alone as enough if the full text contains more specific evidence.
- If a requested field is ambiguous, choose the most conservative answer and mark confidence as "low".
- If evidence is indirect, partial, or only implied, say so in low_confidence_reason.
- If the PDF text is unreadable, incomplete, scanned poorly, or appears to omit pages, mark confidence as "low" and add a review warning.

Anti-miss rules:
- Use an exhaustive extraction strategy for research-question evidence.
- Search the whole article, including abstract, introduction/background, methods, results/findings, discussion, tables, figures, boxes, appendices, and limitations if available.
- For each research question, extract every relevant original-text excerpt you can find.
- Err on the side of including more potentially relevant excerpts rather than fewer.
- Include borderline relevant excerpts when they may help later thematic analysis, and explain the relevance_note.
- Do not omit repeated but meaningfully different evidence from different sections or participant groups.
- Do not collapse multiple distinct findings into one excerpt if separate excerpts would be useful for coding.

Source-location rules:
- Include page numbers when visible or inferable from the PDF.
- If page numbers are not visible, use section names such as "Abstract", "Results", "Discussion", "Table 2", or "Figure 1".
- If neither page nor section is clear, write "location unclear".

Confidence rules:
- Use "high" only when the answer is directly supported by clear article text.
- Use "medium" when the answer is supported but incomplete, scattered, or requires minor interpretation.
- Use "low" when the answer is uncertain, indirect, missing, contradictory, or affected by PDF quality.
- Use low_confidence_reason to explain every medium or low confidence answer in plain language.

Output content rules:
- Return one article record only.
- For every requested structured field, provide a value, source_location, confidence, and low_confidence_reason.
- For every requested research question, provide an answer_summary, confidence, low_confidence_reason, and all relevant excerpts.
- Never leave required values blank. Use "Not found" or "Not relevant to this article" where appropriate.
- Follow the required structured output schema exactly.

Structured fields requested by the user:
{structured_fields}

Research questions requested by the user:
{research_questions}
""".strip()

MMAT_SCREENING_QUESTIONS = [
    {
        "id": "S1",
        "text": "Are there clear research questions?",
    },
    {
        "id": "S2",
        "text": "Do the collected data allow to address the research questions?",
    },
]

MMAT_CATEGORY_CRITERIA = {
    "Qualitative": [
        ("1.1", "Is the qualitative approach appropriate to answer the research question?"),
        ("1.2", "Are the qualitative data collection methods adequate to address the research question?"),
        ("1.3", "Are the findings adequately derived from the data?"),
        ("1.4", "Is the interpretation of results sufficiently substantiated by data?"),
        ("1.5", "Is there coherence between qualitative data sources, collection, analysis and interpretation?"),
    ],
    "Quantitative randomized controlled trial": [
        ("2.1", "Is randomization appropriately performed?"),
        ("2.2", "Are the groups comparable at baseline?"),
        ("2.3", "Are there complete outcome data?"),
        ("2.4", "Are outcome assessors blinded to the intervention provided?"),
        ("2.5", "Did the participants adhere to the assigned intervention?"),
    ],
    "Quantitative non-randomized": [
        ("3.1", "Are the participants representative of the target population?"),
        ("3.2", "Are measurements appropriate regarding both the outcome and intervention or exposure?"),
        ("3.3", "Are there complete outcome data?"),
        ("3.4", "Are the confounders accounted for in the design and analysis?"),
        ("3.5", "During the study period, is the intervention administered or exposure occurred as intended?"),
    ],
    "Quantitative descriptive": [
        ("4.1", "Is the sampling strategy relevant to address the research question?"),
        ("4.2", "Is the sample representative of the target population?"),
        ("4.3", "Are the measurements appropriate?"),
        ("4.4", "Is the risk of nonresponse bias low?"),
        ("4.5", "Is the statistical analysis appropriate to answer the research question?"),
    ],
    "Mixed methods": [
        ("5.1", "Is there an adequate rationale for using a mixed methods design to address the research question?"),
        ("5.2", "Are the different components of the study effectively integrated to answer the research question?"),
        ("5.3", "Are the outputs of the integration of qualitative and quantitative components adequately interpreted?"),
        ("5.4", "Are divergences and inconsistencies between quantitative and qualitative results adequately addressed?"),
        ("5.5", "Do the different components of the study adhere to the quality criteria of each tradition of the methods involved?"),
    ],
}

DEFAULT_MMAT_PROMPT_TEMPLATE = """
You are an expert systematic review quality assessor using the Mixed Methods Appraisal Tool (MMAT), version 2018.

Your task is to read the attached research article PDF and produce one MMAT quality assessment record. Work as a careful reviewer. Do not calculate a total score.

Core rules:
- Use only information found in the attached PDF.
- Do not use outside knowledge or assumptions about similar studies.
- Do not invent methods, study design, sample details, outcome data, page numbers, or author intentions.
- Use the MMAT 2018 response options exactly: "Yes", "No", or "Can't tell".
- If the PDF does not report enough information to answer a criterion, use "Can't tell".
- Give a short plain-language justification for every answer.
- Include page numbers when visible or inferable. If not, use section names such as "Abstract", "Methods", "Results", "Table 1", or "location unclear".
- Mark confidence as "low" when the answer depends on unclear reporting, missing text, poor PDF quality, or difficult study design classification.

MMAT workflow:
1. Answer both screening questions for all PDFs:
   S1. Are there clear research questions?
   S2. Do the collected data allow to address the research questions?
2. Decide whether the paper is an empirical primary study. MMAT is not suitable for reviews, protocols, editorials, commentaries, theoretical papers, and other non-empirical papers.
3. If the paper is suitable for MMAT, classify it into exactly one study design category:
   - Qualitative
   - Quantitative randomized controlled trial
   - Quantitative non-randomized
   - Quantitative descriptive
   - Mixed methods
4. Rate only the five criteria for the chosen category. For mixed methods studies, rate only the five mixed methods criteria; do not expand all qualitative and quantitative criteria.
5. If S1 or S2 is "No" or "Can't tell", continue the assessment if possible, but add a review warning that further MMAT appraisal may not be feasible or appropriate.

Screening questions:
{screening_questions}

MMAT category criteria:
{mmat_criteria}
""".strip()

DEFAULT_EXCLUSION_PROMPT_TEMPLATE = """
You are helping with title and abstract screening for a systematic review.

Your task is to review citation metadata only. You must be very conservative.

Rules:
- Use only the title, abstract, year, journal, DOI, PMID, and authors provided in the JSON input.
- Apply only the exclusion criteria provided by the user.
- Mark ai_suggested_exclusion as true only when the title and/or abstract clearly show the record should be excluded.
- If the evidence is missing, indirect, borderline, or uncertain, mark ai_suggested_exclusion as false.
- Do not exclude records just because the abstract is unavailable.
- The output is only a flag for human review. The user will make the final decision.
- Keep reasons short and plain.

Exclusion criteria:
{exclusion_criteria}

Citation records:
{records_json}
""".strip()


EXTRACTION_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "article": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "title": {"type": "string"},
                "authors": {"type": "string"},
                "year": {"type": "string"},
                "journal": {"type": "string"},
                "overall_confidence": {"type": "string", "enum": CONFIDENCE_LEVELS},
                "low_confidence_reason": {"type": "string"},
            },
            "required": [
                "title",
                "authors",
                "year",
                "journal",
                "overall_confidence",
                "low_confidence_reason",
            ],
        },
        "structured_fields": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "name": {"type": "string"},
                    "value": {"type": "string"},
                    "source_location": {"type": "string"},
                    "confidence": {"type": "string", "enum": CONFIDENCE_LEVELS},
                    "low_confidence_reason": {"type": "string"},
                },
                "required": [
                    "name",
                    "value",
                    "source_location",
                    "confidence",
                    "low_confidence_reason",
                ],
            },
        },
        "research_question_evidence": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "question": {"type": "string"},
                    "answer_summary": {"type": "string"},
                    "confidence": {"type": "string", "enum": CONFIDENCE_LEVELS},
                    "low_confidence_reason": {"type": "string"},
                    "excerpts": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "additionalProperties": False,
                            "properties": {
                                "text": {"type": "string"},
                                "source_location": {"type": "string"},
                                "relevance_note": {"type": "string"},
                            },
                            "required": ["text", "source_location", "relevance_note"],
                        },
                    },
                },
                "required": [
                    "question",
                    "answer_summary",
                    "confidence",
                    "low_confidence_reason",
                    "excerpts",
                ],
            },
        },
        "review_warnings": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
    "required": [
        "article",
        "structured_fields",
        "research_question_evidence",
        "review_warnings",
    ],
}


MMAT_QUESTION_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "criterion_id": {"type": "string"},
        "criterion": {"type": "string"},
        "response": {"type": "string", "enum": MMAT_RESPONSES},
        "justification": {"type": "string"},
        "source_location": {"type": "string"},
        "confidence": {"type": "string", "enum": CONFIDENCE_LEVELS},
        "low_confidence_reason": {"type": "string"},
    },
    "required": [
        "criterion_id",
        "criterion",
        "response",
        "justification",
        "source_location",
        "confidence",
        "low_confidence_reason",
    ],
}


MMAT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "article": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "title": {"type": "string"},
                "authors": {"type": "string"},
                "year": {"type": "string"},
                "journal": {"type": "string"},
            },
            "required": ["title", "authors", "year", "journal"],
        },
        "study_design": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "suitable_for_mmat": {"type": "boolean"},
                "category": {"type": "string", "enum": MMAT_STUDY_DESIGNS},
                "classification_reason": {"type": "string"},
                "needs_human_review": {"type": "boolean"},
            },
            "required": [
                "suitable_for_mmat",
                "category",
                "classification_reason",
                "needs_human_review",
            ],
        },
        "screening_questions": {
            "type": "array",
            "minItems": 2,
            "maxItems": 2,
            "items": MMAT_QUESTION_SCHEMA,
        },
        "category_criteria": {
            "type": "array",
            "minItems": 5,
            "maxItems": 5,
            "items": MMAT_QUESTION_SCHEMA,
        },
        "review_warnings": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
    "required": [
        "article",
        "study_design",
        "screening_questions",
        "category_criteria",
        "review_warnings",
    ],
}


EXCLUSION_MARKING_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "records": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "record_id": {"type": "string"},
                    "ai_suggested_exclusion": {"type": "boolean"},
                    "matched_criteria": {"type": "string"},
                    "reason": {"type": "string"},
                    "evidence": {"type": "string"},
                    "needs_human_review": {"type": "boolean"},
                },
                "required": [
                    "record_id",
                    "ai_suggested_exclusion",
                    "matched_criteria",
                    "reason",
                    "evidence",
                    "needs_human_review",
                ],
            },
        }
    },
    "required": ["records"],
}


def split_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def decode_uploaded_text(uploaded_file: Any) -> str:
    content = uploaded_file.getvalue()
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_whitespace(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_diacritics(value: str) -> str:
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(character)
    )


def normalize_match_text(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"&\s+([a-zA-Z]+)\s*;", r"&\1;", text)
    text = unescape(text)
    text = strip_diacritics(text).casefold()
    text = re.sub(r"(?<=[a-z])(?=\d)|(?<=\d)(?=[a-z])", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_identifier(value: Any) -> str:
    text = normalize_whitespace(value).casefold()
    text = re.sub(r"^(doi|pmid)\s*[:=]\s*", "", text)
    text = text.replace("https://doi.org/", "").replace("http://doi.org/", "")
    text = text.replace("[doi]", "")
    return text.strip().rstrip(".")


def extract_year(value: Any) -> str:
    match = re.search(r"(19|20)\d{2}", str(value or ""))
    return match.group(0) if match else ""


def extract_doi_from_text(value: Any) -> str:
    text = str(value or "")
    match = re.search(r"10\.\d{4,9}/[^\s\]\);,]+", text, flags=re.IGNORECASE)
    return normalize_identifier(match.group(0)) if match else ""


def text_similarity(left: str, right: str) -> float:
    left_norm = normalize_match_text(left)
    right_norm = normalize_match_text(right)
    if not left_norm or not right_norm:
        return 0.0
    return SequenceMatcher(None, left_norm, right_norm).ratio()


def meaningful_tokens(value: Any) -> set[str]:
    stop_words = {
        "a",
        "an",
        "and",
        "are",
        "as",
        "at",
        "by",
        "for",
        "from",
        "in",
        "is",
        "of",
        "on",
        "or",
        "that",
        "the",
        "these",
        "this",
        "to",
        "was",
        "were",
        "with",
    }
    return {
        token
        for token in normalize_match_text(value).split()
        if len(token) > 1 and token not in stop_words
    }


def token_overlap_similarity(left: Any, right: Any) -> float:
    left_tokens = meaningful_tokens(left)
    right_tokens = meaningful_tokens(right)
    if not left_tokens or not right_tokens:
        return 0.0
    return len(left_tokens & right_tokens) / min(len(left_tokens), len(right_tokens))


def parse_tagged_records(text: str, mode: str) -> list[dict[str, list[str]]]:
    records: list[dict[str, list[str]]] = []
    current: dict[str, list[str]] = {}
    current_tag = ""

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n\r")
        if not line.strip():
            continue

        tag = ""
        value = ""
        if mode == "ris":
            match = re.match(r"^([A-Z0-9]{2})\s{2}-\s?(.*)$", line)
        else:
            match = re.match(r"^([A-Z0-9]{2,4})\s*-\s?(.*)$", line)

        if match:
            tag = match.group(1).strip().upper()
            value = match.group(2).strip()
            if tag == "TY" and current:
                records.append(current)
                current = {}
            if mode == "nbib" and tag == "PMID" and current:
                records.append(current)
                current = {}
            current.setdefault(tag, []).append(value)
            current_tag = tag
            if tag == "ER":
                records.append(current)
                current = {}
                current_tag = ""
            continue

        if current_tag and current:
            current[current_tag][-1] = normalize_whitespace(
                f"{current[current_tag][-1]} {line.strip()}"
            )

    if current:
        records.append(current)
    return records


def first_tag(tags: dict[str, list[str]], names: list[str]) -> str:
    for name in names:
        values = tags.get(name, [])
        for value in values:
            text = normalize_whitespace(value)
            if text:
                return text
    return ""


def tag_values(tags: dict[str, list[str]], names: list[str]) -> list[str]:
    values = []
    for name in names:
        for value in tags.get(name, []):
            text = normalize_whitespace(value)
            if text:
                values.append(text)
    return values


def citation_from_tags(
    tags: dict[str, list[str]],
    source_file: str,
    source_format: str,
    record_number: int,
) -> dict[str, Any]:
    if source_format == "nbib":
        doi_candidates = tag_values(tags, ["AID", "LID"]) + tag_values(tags, ["DO"])
        doi = ""
        for candidate in doi_candidates:
            if "[doi]" in candidate.casefold() or extract_doi_from_text(candidate):
                doi = extract_doi_from_text(candidate) or normalize_identifier(candidate)
                doi = doi.replace(" [doi]", "").strip()
                break
        title = first_tag(tags, ["TI", "BTI", "CTI"])
        abstract = " ".join(tag_values(tags, ["AB", "OAB"]))
        journal = first_tag(tags, ["JT", "TA", "JID"])
        year = extract_year(first_tag(tags, ["DP", "DEP", "EDAT", "MHDA"]))
        pmid = normalize_identifier(first_tag(tags, ["PMID"]))
        authors = tag_values(tags, ["AU"]) or tag_values(tags, ["FAU"])
        ris_type = "JOUR"
    else:
        doi_value = first_tag(tags, ["DO"])
        doi = extract_doi_from_text(doi_value) or normalize_identifier(doi_value)
        if not doi:
            doi = extract_doi_from_text(" ".join(tag_values(tags, ["UR", "L1", "L2", "N1"])))
        title = first_tag(tags, ["TI", "T1", "CT", "BT"])
        abstract = " ".join(tag_values(tags, ["AB", "N2"]))
        journal = first_tag(tags, ["JO", "JF", "JA", "T2"])
        year = extract_year(first_tag(tags, ["PY", "Y1", "DA"]))
        pmid = normalize_identifier(first_tag(tags, ["PMID"]))
        if not pmid:
            id_value = first_tag(tags, ["ID"])
            pmid = normalize_identifier(id_value) if re.fullmatch(r"\d+", id_value or "") else ""
        authors = tag_values(tags, ["AU", "A1"])
        ris_type = first_tag(tags, ["TY"]) or "JOUR"

    return {
        "record_id": f"C{record_number:05d}",
        "source_file": source_file,
        "source_format": source_format.upper(),
        "ris_type": ris_type,
        "title": title,
        "abstract": abstract,
        "doi": doi,
        "pmid": pmid,
        "authors": authors,
        "year": year,
        "journal": journal,
        "raw_tags": tags,
        "ai_suggested_exclusion": False,
        "ai_matched_criteria": "",
        "ai_reason": "",
        "ai_evidence": "",
        "needs_human_review": False,
    }


def parse_citation_uploads(
    uploaded_files: list[Any],
) -> tuple[list[dict[str, Any]], list[dict[str, str]], list[dict[str, str]]]:
    records: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    import_log: list[dict[str, str]] = []

    for uploaded_file in uploaded_files:
        filename = uploaded_file.name
        suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
        if suffix not in {"ris", "nbib"}:
            errors.append({"file": filename, "message": "Only .ris and .nbib files are supported."})
            continue

        try:
            text = decode_uploaded_text(uploaded_file)
            tagged_records = parse_tagged_records(text, suffix)
            if not tagged_records:
                errors.append({"file": filename, "message": "No citation records were found."})
                import_log.append(
                    {
                        "source_file": filename,
                        "source_format": suffix.upper(),
                        "parsed_records": "0",
                    }
                )
                continue
            file_record_count = 0
            for tagged_record in tagged_records:
                if suffix == "ris" and not first_tag(tagged_record, ["TY", "TI", "T1"]):
                    continue
                if suffix == "nbib" and not first_tag(tagged_record, ["PMID", "TI"]):
                    continue
                records.append(
                    citation_from_tags(
                        tagged_record,
                        source_file=filename,
                        source_format=suffix,
                        record_number=len(records) + 1,
                    )
                )
                file_record_count += 1
            import_log.append(
                {
                    "source_file": filename,
                    "source_format": suffix.upper(),
                    "parsed_records": str(file_record_count),
                }
            )
        except Exception as exc:
            errors.append({"file": filename, "message": str(exc)})
            import_log.append(
                {
                    "source_file": filename,
                    "source_format": suffix.upper(),
                    "parsed_records": "0",
                }
            )

    return records, errors, import_log


def deduplicate_citations(
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    kept: list[dict[str, Any]] = []
    duplicate_log: list[dict[str, Any]] = []
    doi_index: dict[str, dict[str, Any]] = {}
    pmid_index: dict[str, dict[str, Any]] = {}

    for record in records:
        duplicate_of = None
        reason = ""
        title_similarity = 0.0
        abstract_similarity = 0.0
        abstract_token_overlap = 0.0

        doi = normalize_identifier(record.get("doi"))
        pmid = normalize_identifier(record.get("pmid"))
        if doi and doi in doi_index:
            duplicate_of = doi_index[doi]
            reason = "DOI match"
        elif pmid and pmid in pmid_index:
            duplicate_of = pmid_index[pmid]
            reason = "PMID match"
        else:
            for kept_record in kept:
                kept_doi = normalize_identifier(kept_record.get("doi"))
                kept_pmid = normalize_identifier(kept_record.get("pmid"))
                if doi and kept_doi and doi != kept_doi:
                    continue
                if pmid and kept_pmid and pmid != kept_pmid:
                    continue
                title_similarity = text_similarity(record.get("title", ""), kept_record.get("title", ""))
                abstract_similarity = text_similarity(
                    record.get("abstract", ""), kept_record.get("abstract", "")
                )
                abstract_token_overlap = token_overlap_similarity(
                    record.get("abstract", ""),
                    kept_record.get("abstract", ""),
                )
                if (
                    title_similarity >= SIMILARITY_THRESHOLD
                    and (
                        abstract_similarity >= SIMILARITY_THRESHOLD
                        or abstract_token_overlap >= CITATION_ABSTRACT_TOKEN_OVERLAP_THRESHOLD
                    )
                ):
                    duplicate_of = kept_record
                    reason = "Title similarity >= 95% and abstract similarity/token overlap >= 95%"
                    break

        if duplicate_of:
            duplicate_log.append(
                {
                    "removed_record_id": record.get("record_id", ""),
                    "removed_title": record.get("title", ""),
                    "removed_doi": record.get("doi", ""),
                    "removed_pmid": record.get("pmid", ""),
                    "removed_source_file": record.get("source_file", ""),
                    "kept_record_id": duplicate_of.get("record_id", ""),
                    "kept_title": duplicate_of.get("title", ""),
                    "kept_doi": duplicate_of.get("doi", ""),
                    "kept_pmid": duplicate_of.get("pmid", ""),
                    "kept_source_file": duplicate_of.get("source_file", ""),
                    "duplicate_reason": reason,
                    "title_similarity": f"{title_similarity:.3f}" if title_similarity else "",
                    "abstract_similarity": f"{abstract_similarity:.3f}" if abstract_similarity else "",
                    "abstract_token_overlap": f"{abstract_token_overlap:.3f}" if abstract_token_overlap else "",
                }
            )
            continue

        kept.append(record)
        if doi:
            doi_index[doi] = record
        if pmid:
            pmid_index[pmid] = record

    return kept, duplicate_log


def citation_to_ai_payload(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "record_id": record.get("record_id", ""),
        "title": normalize_whitespace(record.get("title", "")),
        "abstract": normalize_whitespace(record.get("abstract", "")),
    }


def make_exclusion_prompt(
    records: list[dict[str, Any]],
    exclusion_criteria: list[str],
    prompt_template: str,
) -> str:
    criteria_text = "\n".join(f"- {criterion}" for criterion in exclusion_criteria)
    records_json = json.dumps(
        [citation_to_ai_payload(record) for record in records],
        ensure_ascii=False,
        separators=(",", ":"),
    )
    template = prompt_template.strip() or DEFAULT_EXCLUSION_PROMPT_TEMPLATE
    if "{exclusion_criteria}" not in template:
        template = f"{template}\n\nExclusion criteria:\n{{exclusion_criteria}}"
    if "{records_json}" not in template:
        template = f"{template}\n\nCitation records:\n{{records_json}}"
    return (
        template.replace("{exclusion_criteria}", criteria_text)
        .replace("{records_json}", records_json)
        .strip()
    )


def mark_citation_exclusions(
    records: list[dict[str, Any]],
    exclusion_criteria: list[str],
    api_key: str,
    base_url: str,
    model: str,
    prompt_template: str,
) -> tuple[list[dict[str, Any]], str]:
    client = OpenAI(api_key=api_key, base_url=base_url.rstrip("/"))
    prompt = make_exclusion_prompt(records, exclusion_criteria, prompt_template)

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [{"type": "input_text", "text": prompt}],
            }
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "citation_exclusion_marking",
                "strict": True,
                "schema": EXCLUSION_MARKING_SCHEMA,
            }
        },
    )

    data = json.loads(response.output_text)
    decisions = {
        clean_text(item.get("record_id"), ""): item
        for item in data.get("records", [])
        if isinstance(item, dict)
    }
    for record in records:
        decision = decisions.get(record.get("record_id", ""), {})
        record["ai_suggested_exclusion"] = clean_bool(
            decision.get("ai_suggested_exclusion"),
            False,
        )
        record["ai_matched_criteria"] = clean_text(decision.get("matched_criteria"), "")
        record["ai_reason"] = clean_text(decision.get("reason"), "")
        record["ai_evidence"] = clean_text(decision.get("evidence"), "")
        record["needs_human_review"] = clean_bool(decision.get("needs_human_review"), False)
    return records, prompt


def citation_ai_payload_size(record: dict[str, Any]) -> int:
    payload = citation_to_ai_payload(record)
    return len(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))


def batch_citations_for_ai(records: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    batches: list[list[dict[str, Any]]] = []
    current_batch: list[dict[str, Any]] = []
    current_size = 0

    for record in records:
        record_size = citation_ai_payload_size(record)
        would_exceed_records = len(current_batch) >= CITATION_AI_MAX_BATCH_RECORDS
        would_exceed_chars = (
            current_batch
            and current_size + record_size > CITATION_AI_BATCH_CHAR_BUDGET
        )
        if would_exceed_records or would_exceed_chars:
            batches.append(current_batch)
            current_batch = []
            current_size = 0

        current_batch.append(record)
        current_size += record_size

    if current_batch:
        batches.append(current_batch)
    return batches


def citation_exclusion_prompt_note(
    prompt_template: str,
    exclusion_criteria: list[str],
    total_records: int,
) -> str:
    criteria_text = "\n".join(f"- {criterion}" for criterion in exclusion_criteria)
    return "\n\n".join(
        [
            "AI citation exclusion marking was run in batches to avoid token-per-minute limits.",
            f"Maximum batch size: {CITATION_AI_MAX_BATCH_RECORDS} records",
            f"Approximate batch character budget: {CITATION_AI_BATCH_CHAR_BUDGET}",
            "AI input fields: original title and original abstract only",
            "No title or abstract text was truncated for AI marking.",
            f"Total records sent for AI marking: {total_records}",
            "Exclusion criteria:",
            criteria_text or "not provided",
            "Prompt template used:",
            prompt_template.strip() or DEFAULT_EXCLUSION_PROMPT_TEMPLATE,
        ]
    )


def mark_citation_exclusions_batched(
    records: list[dict[str, Any]],
    exclusion_criteria: list[str],
    api_key: str,
    base_url: str,
    model: str,
    prompt_template: str,
    status: Any | None = None,
    progress: Any | None = None,
) -> tuple[list[dict[str, Any]], str]:
    if not records:
        return records, citation_exclusion_prompt_note(prompt_template, exclusion_criteria, 0)

    batches = batch_citations_for_ai(records)
    marked_records: list[dict[str, Any]] = []
    total_batches = len(batches)

    for batch_index, batch in enumerate(batches, start=1):
        if status is not None:
            status.info(
                f"AI marking citation batch {batch_index}/{total_batches} "
                f"({len(batch)} records)"
            )
        marked_batch, _prompt_used = mark_citation_exclusions(
            records=batch,
            exclusion_criteria=exclusion_criteria,
            api_key=api_key,
            base_url=base_url,
            model=model,
            prompt_template=prompt_template,
        )
        marked_records.extend(marked_batch)
        if progress is not None:
            progress.progress(batch_index / total_batches)
        if batch_index < total_batches and CITATION_AI_BATCH_DELAY_SECONDS > 0:
            sleep(CITATION_AI_BATCH_DELAY_SECONDS)

    return marked_records, citation_exclusion_prompt_note(
        prompt_template,
        exclusion_criteria,
        len(records),
    )


def make_prompt(fields: list[str], questions: list[str], prompt_template: str) -> str:
    field_list = "\n".join(f"- {field}" for field in fields) or "- No extra structured fields"
    question_list = "\n".join(
        f"- RQ{index}: {question}" for index, question in enumerate(questions, start=1)
    ) or "- No research questions"
    template = prompt_template.strip() or DEFAULT_PROMPT_TEMPLATE
    if "{structured_fields}" not in template:
        template = f"{template}\n\nStructured fields requested by the user:\n{{structured_fields}}"
    if "{research_questions}" not in template:
        template = f"{template}\n\nResearch questions requested by the user:\n{{research_questions}}"
    return (
        template.replace("{structured_fields}", field_list)
        .replace("{research_questions}", question_list)
        .strip()
    )


def format_mmat_criteria() -> str:
    sections = []
    for category, criteria in MMAT_CATEGORY_CRITERIA.items():
        lines = [f"{category}:"]
        lines.extend(f"- {criterion_id}. {criterion}" for criterion_id, criterion in criteria)
        sections.append("\n".join(lines))
    return "\n\n".join(sections)


def make_mmat_prompt(prompt_template: str) -> str:
    screening_list = "\n".join(
        f"- {item['id']}. {item['text']}" for item in MMAT_SCREENING_QUESTIONS
    )
    template = prompt_template.strip() or DEFAULT_MMAT_PROMPT_TEMPLATE
    if "{screening_questions}" not in template:
        template = f"{template}\n\nScreening questions:\n{{screening_questions}}"
    if "{mmat_criteria}" not in template:
        template = f"{template}\n\nMMAT category criteria:\n{{mmat_criteria}}"
    return (
        template.replace("{screening_questions}", screening_list)
        .replace("{mmat_criteria}", format_mmat_criteria())
        .strip()
    )


def pdf_to_input_file(uploaded_file: Any) -> dict[str, str]:
    encoded = base64.b64encode(uploaded_file.getvalue()).decode("utf-8")
    return {
        "type": "input_file",
        "filename": uploaded_file.name,
        "file_data": f"data:application/pdf;base64,{encoded}",
    }


def extract_from_pdf(
    uploaded_file: Any,
    api_key: str,
    base_url: str,
    model: str,
    fields: list[str],
    questions: list[str],
    prompt_template: str,
) -> dict[str, Any]:
    client = OpenAI(api_key=api_key, base_url=base_url.rstrip("/"))
    prompt = make_prompt(fields, questions, prompt_template)

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    pdf_to_input_file(uploaded_file),
                    {"type": "input_text", "text": prompt},
                ],
            }
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "systematic_review_extraction",
                "strict": True,
                "schema": EXTRACTION_SCHEMA,
            }
        },
    )

    raw_text = response.output_text
    data = json.loads(raw_text)
    data = normalize_extraction_result(data, fields, questions)
    data["source_file"] = uploaded_file.name
    data["requested_fields"] = fields
    data["requested_questions"] = questions
    data["prompt_used"] = prompt
    return data


def assess_quality_from_pdf(
    uploaded_file: Any,
    api_key: str,
    base_url: str,
    model: str,
    prompt_template: str,
) -> dict[str, Any]:
    client = OpenAI(api_key=api_key, base_url=base_url.rstrip("/"))
    prompt = make_mmat_prompt(prompt_template)

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    pdf_to_input_file(uploaded_file),
                    {"type": "input_text", "text": prompt},
                ],
            }
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "mmat_quality_assessment",
                "strict": True,
                "schema": MMAT_SCHEMA,
            }
        },
    )

    raw_text = response.output_text
    data = json.loads(raw_text)
    data = normalize_mmat_result(data)
    data["source_file"] = uploaded_file.name
    data["mmat_prompt_used"] = prompt
    return data


def clean_text(value: Any, default: str = "Not found") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def clean_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().casefold() in {"true", "yes", "1"}
    return default


def clean_confidence(value: Any) -> str:
    confidence = clean_text(value, "low").lower()
    return confidence if confidence in CONFIDENCE_LEVELS else "low"


def clean_mmat_response(value: Any) -> str:
    response = clean_text(value, "Can't tell")
    return response if response in MMAT_RESPONSES else "Can't tell"


def normalize_mmat_question(item: dict[str, Any], criterion_id: str, criterion: str) -> dict[str, str]:
    return {
        "criterion_id": clean_text(item.get("criterion_id"), criterion_id),
        "criterion": clean_text(item.get("criterion"), criterion),
        "response": clean_mmat_response(item.get("response")),
        "justification": clean_text(item.get("justification")),
        "source_location": clean_text(item.get("source_location")),
        "confidence": clean_confidence(item.get("confidence")),
        "low_confidence_reason": clean_text(item.get("low_confidence_reason"), ""),
    }


def normalize_mmat_result(data: dict[str, Any]) -> dict[str, Any]:
    article = data.setdefault("article", {})
    for key in ("title", "authors", "year", "journal"):
        article[key] = clean_text(article.get(key))

    study_design = data.setdefault("study_design", {})
    category = clean_text(study_design.get("category"), "Not suitable for MMAT")
    if category not in MMAT_STUDY_DESIGNS:
        category = "Not suitable for MMAT"
    study_design["category"] = category
    study_design["suitable_for_mmat"] = clean_bool(
        study_design.get("suitable_for_mmat"),
        category != "Not suitable for MMAT",
    )
    study_design["classification_reason"] = clean_text(
        study_design.get("classification_reason")
    )
    study_design["needs_human_review"] = clean_bool(
        study_design.get("needs_human_review"),
        category == "Not suitable for MMAT",
    )

    returned_screening = [
        item for item in data.get("screening_questions", []) if isinstance(item, dict)
    ]
    normalized_screening = []
    for index, expected in enumerate(MMAT_SCREENING_QUESTIONS):
        item = returned_screening[index] if index < len(returned_screening) else {}
        normalized_screening.append(
            normalize_mmat_question(item, expected["id"], expected["text"])
        )
    data["screening_questions"] = normalized_screening

    expected_criteria = MMAT_CATEGORY_CRITERIA.get(category)
    if not expected_criteria:
        expected_criteria = [
            ("N/A1", "MMAT category criterion is not applicable because the paper is not suitable for MMAT."),
            ("N/A2", "MMAT category criterion is not applicable because the paper is not suitable for MMAT."),
            ("N/A3", "MMAT category criterion is not applicable because the paper is not suitable for MMAT."),
            ("N/A4", "MMAT category criterion is not applicable because the paper is not suitable for MMAT."),
            ("N/A5", "MMAT category criterion is not applicable because the paper is not suitable for MMAT."),
        ]
    returned_criteria = [
        item for item in data.get("category_criteria", []) if isinstance(item, dict)
    ]
    normalized_criteria = []
    for index, (criterion_id, criterion) in enumerate(expected_criteria):
        item = returned_criteria[index] if index < len(returned_criteria) else {}
        normalized_criteria.append(normalize_mmat_question(item, criterion_id, criterion))
    data["category_criteria"] = normalized_criteria

    warnings = [
        clean_text(warning, "")
        for warning in data.get("review_warnings", [])
        if clean_text(warning, "")
    ]
    if any(item["response"] != "Yes" for item in normalized_screening):
        warnings.append(
            "One or both MMAT screening questions were not answered Yes; further appraisal may not be feasible or appropriate."
        )
    if not study_design["suitable_for_mmat"]:
        warnings.append("This paper was marked as not suitable for MMAT appraisal.")
    data["review_warnings"] = list(dict.fromkeys(warnings))
    return data


def normalize_extraction_result(
    data: dict[str, Any],
    fields: list[str],
    questions: list[str],
) -> dict[str, Any]:
    article = data.setdefault("article", {})
    for key in ("title", "authors", "year", "journal"):
        article[key] = clean_text(article.get(key))
    article["overall_confidence"] = clean_text(article.get("overall_confidence"), "low").lower()
    if article["overall_confidence"] not in CONFIDENCE_LEVELS:
        article["overall_confidence"] = "low"
    article["low_confidence_reason"] = clean_text(article.get("low_confidence_reason"))

    returned_fields = {
        clean_text(item.get("name"), "").casefold(): item
        for item in data.get("structured_fields", [])
        if isinstance(item, dict)
    }
    normalized_fields = []
    for field in fields:
        item = returned_fields.get(field.casefold(), {})
        confidence = clean_text(item.get("confidence"), "low").lower()
        normalized_fields.append(
            {
                "name": field,
                "value": clean_text(item.get("value")),
                "source_location": clean_text(item.get("source_location")),
                "confidence": confidence if confidence in CONFIDENCE_LEVELS else "low",
                "low_confidence_reason": clean_text(item.get("low_confidence_reason")),
            }
        )
    data["structured_fields"] = normalized_fields

    returned_evidence = [
        item for item in data.get("research_question_evidence", []) if isinstance(item, dict)
    ]
    normalized_evidence = []
    for index, question in enumerate(questions, start=1):
        item = returned_evidence[index - 1] if index - 1 < len(returned_evidence) else {}
        confidence = clean_text(item.get("confidence"), "low").lower()
        excerpts = []
        for excerpt in item.get("excerpts", []) or []:
            if not isinstance(excerpt, dict):
                continue
            excerpts.append(
                {
                    "text": clean_text(excerpt.get("text")),
                    "source_location": clean_text(excerpt.get("source_location")),
                    "relevance_note": clean_text(excerpt.get("relevance_note")),
                }
            )
        normalized_evidence.append(
            {
                "question": f"RQ{index}: {question}",
                "answer_summary": clean_text(item.get("answer_summary")),
                "confidence": confidence if confidence in CONFIDENCE_LEVELS else "low",
                "low_confidence_reason": clean_text(item.get("low_confidence_reason")),
                "excerpts": excerpts,
            }
        )
    data["research_question_evidence"] = normalized_evidence
    data["review_warnings"] = [
        clean_text(warning, "") for warning in data.get("review_warnings", []) if clean_text(warning, "")
    ]
    return data


def requested_fields_for_result(result: dict[str, Any]) -> list[str]:
    fields = result.get("requested_fields")
    if fields:
        return list(fields)
    return [item.get("name", "") for item in result.get("structured_fields", []) if item.get("name")]


def requested_questions_for_result(result: dict[str, Any]) -> list[str]:
    questions = result.get("requested_questions")
    if questions:
        return list(questions)
    return [
        evidence.get("question", "")
        for evidence in result.get("research_question_evidence", [])
        if evidence.get("question")
    ]


def result_to_flat_row(result: dict[str, Any]) -> dict[str, str]:
    article = result.get("article", {})
    row = {
        "File names": result.get("source_file", ""),
        "title": article.get("title", ""),
        "authors": article.get("authors", ""),
        "year": article.get("year", ""),
        "journal": article.get("journal", ""),
        "overall_confidence": article.get("overall_confidence", ""),
        "review_warnings": "; ".join(result.get("review_warnings", [])),
    }

    fields_by_name = {
        item.get("name", "").casefold(): item for item in result.get("structured_fields", [])
    }
    for field in requested_fields_for_result(result):
        item = fields_by_name.get(field.casefold(), {})
        row[field] = clean_text(item.get("value"))
        row[f"{field} confidence"] = clean_text(item.get("confidence"), "low")

    evidence_items = result.get("research_question_evidence", [])
    for index, _question in enumerate(requested_questions_for_result(result), start=1):
        evidence = evidence_items[index - 1] if index - 1 < len(evidence_items) else {}
        row[f"RQ{index} summary"] = clean_text(evidence.get("answer_summary"))
        row[f"RQ{index} confidence"] = clean_text(evidence.get("confidence"), "low")

    return row


def confidence_needs_review(value: str) -> bool:
    return value.lower() in {"low", "medium"}


def mmat_response_needs_review(value: str) -> bool:
    return value in {"No", "Can't tell"}


def style_results(df: pd.DataFrame) -> Any:
    def highlight(value: Any) -> str:
        if isinstance(value, str) and (
            confidence_needs_review(value) or mmat_response_needs_review(value)
        ):
            return "background-color: #ffe5e5; color: #7a1f1f;"
        return ""

    return df.style.map(highlight)


def result_to_evidence_rows(result: dict[str, Any]) -> list[dict[str, str]]:
    article = result.get("article", {})
    rows = []
    for question_index, evidence in enumerate(result.get("research_question_evidence", []), start=1):
        excerpts = evidence.get("excerpts", []) or [{"text": "Not found", "source_location": "Not found", "relevance_note": "Not found"}]
        for excerpt in excerpts:
            rows.append(
                {
                    "File names": result.get("source_file", ""),
                    "title": clean_text(article.get("title")),
                    "research question": f"RQ{question_index}",
                    "answer_summary": clean_text(evidence.get("answer_summary")),
                    "confidence": clean_text(evidence.get("confidence"), "low"),
                    "excerpt": clean_text(excerpt.get("text")),
                    "source_location": clean_text(excerpt.get("source_location")),
                    "relevance_note": clean_text(excerpt.get("relevance_note")),
                }
            )
    return rows


def mmat_result_to_summary_row(result: dict[str, Any]) -> dict[str, str]:
    article = result.get("article", {})
    study_design = result.get("study_design", {})
    row = {
        "File names": result.get("source_file", ""),
        "title": clean_text(article.get("title")),
        "authors": clean_text(article.get("authors")),
        "year": clean_text(article.get("year")),
        "journal": clean_text(article.get("journal")),
        "suitable_for_mmat": str(study_design.get("suitable_for_mmat", False)),
        "study_design_category": clean_text(study_design.get("category")),
        "classification_reason": clean_text(study_design.get("classification_reason")),
        "needs_human_review": str(study_design.get("needs_human_review", False)),
        "review_warnings": "; ".join(result.get("review_warnings", [])),
    }

    for item in result.get("screening_questions", []):
        criterion_id = clean_text(item.get("criterion_id"), "S")
        row[f"{criterion_id} response"] = clean_text(item.get("response"), "Can't tell")
        row[f"{criterion_id} confidence"] = clean_text(item.get("confidence"), "low")

    for item in result.get("category_criteria", []):
        criterion_id = clean_text(item.get("criterion_id"), "criterion")
        row[f"{criterion_id} response"] = clean_text(item.get("response"), "Can't tell")
        row[f"{criterion_id} confidence"] = clean_text(item.get("confidence"), "low")

    return row


def mmat_result_to_evidence_rows(result: dict[str, Any]) -> list[dict[str, str]]:
    article = result.get("article", {})
    study_design = result.get("study_design", {})
    rows = []
    for section, items in (
        ("Screening", result.get("screening_questions", [])),
        ("Category criteria", result.get("category_criteria", [])),
    ):
        for item in items:
            rows.append(
                {
                    "File names": result.get("source_file", ""),
                    "title": clean_text(article.get("title")),
                    "study_design_category": clean_text(study_design.get("category")),
                    "section": section,
                    "criterion_id": clean_text(item.get("criterion_id")),
                    "criterion": clean_text(item.get("criterion")),
                    "response": clean_text(item.get("response"), "Can't tell"),
                    "justification": clean_text(item.get("justification")),
                    "source_location": clean_text(item.get("source_location")),
                    "confidence": clean_text(item.get("confidence"), "low"),
                    "low_confidence_reason": clean_text(item.get("low_confidence_reason"), ""),
                }
            )
    return rows


def citation_to_screening_row(record: dict[str, Any]) -> dict[str, str]:
    return {
        "record_id": clean_text(record.get("record_id"), ""),
        "title": clean_text(record.get("title"), ""),
        "abstract": clean_text(record.get("abstract"), ""),
        "doi": clean_text(record.get("doi"), ""),
        "pmid": clean_text(record.get("pmid"), ""),
        "authors": "; ".join(record.get("authors", [])),
        "year": clean_text(record.get("year"), ""),
        "journal": clean_text(record.get("journal"), ""),
        "source_file": clean_text(record.get("source_file"), ""),
        "source_format": clean_text(record.get("source_format"), ""),
        "AI suggested exclusion": "Yes" if record.get("ai_suggested_exclusion") else "No",
        "matched exclusion criteria": clean_text(record.get("ai_matched_criteria"), ""),
        "AI reason": clean_text(record.get("ai_reason"), ""),
        "AI evidence": clean_text(record.get("ai_evidence"), ""),
        "needs human review": "Yes" if record.get("needs_human_review") else "No",
    }


def build_screening_excel_export(
    records: list[dict[str, Any]],
    duplicate_log: list[dict[str, Any]],
    import_log: list[dict[str, Any]],
    exclusion_criteria: list[str],
    ai_prompt_used: str,
) -> bytes:
    workbook = Workbook()
    results_sheet = workbook.active
    results_sheet.title = "Screening Results"
    screening_rows = [citation_to_screening_row(record) for record in records]
    if screening_rows:
        add_rows_to_sheet(results_sheet, screening_rows)
    else:
        results_sheet.append(
            [
                "record_id",
                "title",
                "abstract",
                "doi",
                "pmid",
                "authors",
                "year",
                "journal",
                "AI suggested exclusion",
            ]
        )
    tune_excel_sheet(results_sheet)

    duplicate_sheet = workbook.create_sheet("Duplicate Log")
    if duplicate_log:
        add_rows_to_sheet(duplicate_sheet, duplicate_log)
    else:
        duplicate_sheet.append(
            [
                "removed_record_id",
                "removed_title",
                "kept_record_id",
                "kept_title",
                "duplicate_reason",
            ]
        )
    tune_excel_sheet(duplicate_sheet)

    import_sheet = workbook.create_sheet("Import Log")
    if import_log:
        add_rows_to_sheet(import_sheet, import_log)
    else:
        import_sheet.append(["source_file", "source_format", "parsed_records"])
    tune_excel_sheet(import_sheet)

    methodology_sheet = workbook.create_sheet("Methodology Criteria")
    methodology_sheet.append(["item", "value"])
    methodology_sheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    methodology_sheet.append(["Deduplication rule 1", "If DOI or PMID matches, the later record is removed as duplicate."])
    methodology_sheet.append(["Deduplication rule 2", "If DOI and PMID are missing, title similarity >= 95% and abstract similarity >= 95% removes the later record as duplicate."])
    methodology_sheet.append(["Exclusion criteria", "\n".join(exclusion_criteria) if exclusion_criteria else "not provided"])
    methodology_sheet.append(["AI marking prompt used", ai_prompt_used or "not run"])
    methodology_sheet.append(["AI deletion note", "AI-marked irrelevant records are not deleted from the screening Excel."])
    tune_excel_sheet(methodology_sheet)
    methodology_sheet.column_dimensions["A"].width = 26
    methodology_sheet.column_dimensions["B"].width = 100
    methodology_sheet.row_dimensions[6].height = 240

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def ris_clean(value: Any) -> str:
    return normalize_whitespace(value).replace("\r", " ").replace("\n", " ")


def ris_lines_for_record(record: dict[str, Any]) -> list[str]:
    lines = [f"TY  - {ris_clean(record.get('ris_type') or 'JOUR')}"]
    for author in record.get("authors", []):
        author_text = ris_clean(author)
        if author_text:
            lines.append(f"AU  - {author_text}")
    field_map = [
        ("TI", record.get("title")),
        ("AB", record.get("abstract")),
        ("DO", record.get("doi")),
        ("ID", record.get("pmid")),
        ("PY", record.get("year")),
        ("JO", record.get("journal")),
    ]
    for tag, value in field_map:
        text = ris_clean(value)
        if text:
            lines.append(f"{tag}  - {text}")
    lines.append("ER  -")
    return lines


def build_ris_export(records: list[dict[str, Any]]) -> bytes:
    lines: list[str] = []
    for record in records:
        lines.extend(ris_lines_for_record(record))
        lines.append("")
    return "\n".join(lines).encode("utf-8")


def add_rows_to_sheet(sheet: Any, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def tune_excel_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill("solid", fgColor="FCE4E4")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    header_by_column = {
        cell.column: str(cell.value or "").casefold()
        for cell in sheet[1]
    }

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            is_confidence_cell = "confidence" in header_by_column.get(cell.column, "")
            is_response_cell = "response" in header_by_column.get(cell.column, "")
            if (
                isinstance(cell.value, str)
                and (
                    (is_confidence_cell and confidence_needs_review(cell.value))
                    or (is_response_cell and mmat_response_needs_review(cell.value))
                )
            ):
                cell.fill = review_fill

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        longest_line = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines() or [""]:
                longest_line = max(longest_line, len(line))
        sheet.column_dimensions[column_letter].width = min(max(longest_line + 2, 12), 60)

    for row in sheet.iter_rows():
        max_lines = 1
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            wrapped_lines = max(1, len(value) // 55 + 1)
            explicit_lines = value.count("\n") + 1
            max_lines = max(max_lines, wrapped_lines, explicit_lines)
        sheet.row_dimensions[row[0].row].height = min(max(18, max_lines * 15), 120)


def merge_repeated_evidence_cells(sheet: Any) -> None:
    if sheet.max_row < 3:
        return

    merge_columns = [1, 2, 3, 4, 5]
    group_start = 2
    previous_key = None

    for row_index in range(2, sheet.max_row + 2):
        if row_index <= sheet.max_row:
            current_key = (
                sheet.cell(row=row_index, column=1).value,
                sheet.cell(row=row_index, column=2).value,
                sheet.cell(row=row_index, column=3).value,
            )
        else:
            current_key = None

        if previous_key is None:
            previous_key = current_key
            continue

        if current_key != previous_key:
            group_end = row_index - 1
            if group_end > group_start:
                for column_index in merge_columns:
                    sheet.merge_cells(
                        start_row=group_start,
                        start_column=column_index,
                        end_row=group_end,
                        end_column=column_index,
                    )
                    sheet.cell(row=group_start, column=column_index).alignment = Alignment(
                        vertical="center",
                        wrap_text=True,
                    )
            group_start = row_index
            previous_key = current_key


def build_excel_export(
    results: list[dict[str, Any]],
    qa_results: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Article Summary"
    summary_sheet["A1"] = "No extraction results"

    summary_rows = [result_to_flat_row(result) for result in results]
    evidence_rows = []
    for result in results:
        evidence_rows.extend(result_to_evidence_rows(result))

    if summary_rows:
        summary_sheet.delete_rows(1, 1)
        add_rows_to_sheet(summary_sheet, summary_rows)
        tune_excel_sheet(summary_sheet)

    evidence_sheet = workbook.create_sheet("Evidence Excerpts")
    if evidence_rows:
        add_rows_to_sheet(evidence_sheet, evidence_rows)
    else:
        evidence_sheet.append(["File names", "title", "research question", "answer_summary", "confidence", "excerpt", "source_location", "relevance_note"])
    merge_repeated_evidence_cells(evidence_sheet)
    tune_excel_sheet(evidence_sheet)

    mmat_summary_sheet = workbook.create_sheet("MMAT Summary")
    mmat_summary_rows = [mmat_result_to_summary_row(result) for result in qa_results]
    if mmat_summary_rows:
        add_rows_to_sheet(mmat_summary_sheet, mmat_summary_rows)
    else:
        mmat_summary_sheet.append(
            [
                "File names",
                "title",
                "suitable_for_mmat",
                "study_design_category",
                "S1 response",
                "S2 response",
                "review_warnings",
            ]
        )
    tune_excel_sheet(mmat_summary_sheet)

    mmat_evidence_sheet = workbook.create_sheet("MMAT Evidence")
    mmat_evidence_rows = []
    for result in qa_results:
        mmat_evidence_rows.extend(mmat_result_to_evidence_rows(result))
    if mmat_evidence_rows:
        add_rows_to_sheet(mmat_evidence_sheet, mmat_evidence_rows)
    else:
        mmat_evidence_sheet.append(
            [
                "File names",
                "title",
                "study_design_category",
                "section",
                "criterion_id",
                "criterion",
                "response",
                "justification",
                "source_location",
                "confidence",
                "low_confidence_reason",
            ]
        )
    tune_excel_sheet(mmat_evidence_sheet)

    methodology_sheet = workbook.create_sheet("Methodology Prompt")
    methodology_sheet.append(["item", "value"])
    methodology_sheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    methodology_sheet.append(["Extraction prompt used", results[0].get("prompt_used", "not recorded") if results else "not recorded"])
    methodology_sheet.append(["MMAT prompt used", qa_results[0].get("mmat_prompt_used", "not recorded") if qa_results else "not recorded"])
    methodology_sheet.append(["Prompt note", "These are the actual prompt texts sent to the AI model for extraction and MMAT quality assessment."])
    tune_excel_sheet(methodology_sheet)
    methodology_sheet.column_dimensions["A"].width = 22
    methodology_sheet.column_dimensions["B"].width = 100
    methodology_sheet.row_dimensions[3].height = 240
    methodology_sheet.row_dimensions[4].height = 240

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def initialise_state() -> None:
    for stale_key in (
        "prompt_template_editor",
        "mmat_prompt_template_editor",
        "citation_exclusion_prompt_editor",
    ):
        st.session_state.pop(stale_key, None)
    st.session_state.setdefault("results", [])
    st.session_state.setdefault("errors", [])
    st.session_state.setdefault("qa_results", [])
    st.session_state.setdefault("qa_errors", [])
    st.session_state.setdefault("citation_records", [])
    st.session_state.setdefault("citation_duplicate_log", [])
    st.session_state.setdefault("citation_import_log", [])
    st.session_state.setdefault("citation_errors", [])
    st.session_state.setdefault("citation_ai_prompt_used", "")
    st.session_state.setdefault("citation_imported_count", 0)
    st.session_state.setdefault("citation_export_timestamp", "")
    st.session_state.setdefault("citation_exclusion_prompt_template", DEFAULT_EXCLUSION_PROMPT_TEMPLATE)
    st.session_state.setdefault("research_questions", [""])
    st.session_state.setdefault("prompt_template", DEFAULT_PROMPT_TEMPLATE)
    st.session_state.setdefault("mmat_prompt_template", DEFAULT_MMAT_PROMPT_TEMPLATE)


def apply_custom_style() -> None:
    st.markdown(
        """
        <style>
        :root {
            --app-bg: #f6f8fb;
            --panel-bg: #ffffff;
            --sidebar-bg: #eef2f7;
            --text-main: #111827;
            --text-muted: #64748b;
            --border: #dfe6ef;
            --border-strong: #c7d2df;
            --accent: #334155;
            --accent-soft: #eef3f8;
            --shadow: 0 18px 46px rgba(15, 23, 42, 0.08);
        }

        html, body, [class*="css"] {
            font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display", "SF Pro Text", Inter, "Segoe UI", sans-serif;
        }

        .stApp {
            background: var(--app-bg);
            color: var(--text-main);
        }

        [data-testid="stSidebar"] {
            position: relative;
            background: var(--sidebar-bg);
            border-right: 1px solid var(--border);
        }

        [data-testid="stHeader"],
        header[data-testid="stHeader"],
        [data-testid="stAppViewContainer"] > header,
        .stAppHeader {
            background: var(--app-bg) !important;
            box-shadow: none !important;
            border-bottom: 1px solid var(--border) !important;
        }

        [data-testid="stToolbar"],
        .stAppToolbar {
            background: transparent !important;
        }

        #MainMenu,
        .stDeployButton,
        .stAppDeployButton {
            display: none !important;
            visibility: hidden !important;
        }

        .top-brand {
            position: fixed;
            top: 1.28rem;
            left: 4.1rem;
            z-index: 999990;
            color: var(--text-main);
            font-size: 1.24rem;
            font-weight: 760;
            letter-spacing: 0;
            line-height: 1;
            pointer-events: none;
        }

        .sidebar-brand {
            position: relative;
            top: -5.05rem;
            color: var(--text-main);
            font-size: 1.24rem;
            font-weight: 760;
            line-height: 1;
            margin: 0 0 -2.35rem;
            letter-spacing: 0;
        }

        [data-testid="stSidebar"] > div:first-child {
            padding-top: 1.15rem;
            padding-left: 1.15rem;
            padding-right: 1.15rem;
        }

        [data-testid="stSidebar"] button.e7msn5c15 {
            position: absolute !important;
            top: 0.83rem !important;
            left: 15.6rem !important;
            z-index: 3 !important;
            color: var(--text-muted) !important;
            opacity: 1 !important;
            visibility: visible !important;
        }

        [data-testid="stSidebar"] button.e7msn5c15 span {
            color: var(--text-muted) !important;
        }

        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3 {
            color: var(--text-main);
            letter-spacing: 0;
            font-size: 1rem;
            font-weight: 680;
        }

        .main .block-container {
            max-width: 1120px;
            padding-top: 2.15rem;
            padding-bottom: 2.8rem;
        }

        h1 {
            color: var(--text-main);
            font-weight: 650;
            font-size: clamp(2.15rem, 4vw, 3.35rem);
            letter-spacing: 0;
            line-height: 1.04;
            margin-bottom: 0.45rem;
        }

        div[data-testid="stCaptionContainer"] {
            color: var(--text-muted);
        }

        label, .stMarkdown p {
            color: var(--text-main);
        }

        .stTextInput input,
        .stTextArea textarea {
            border-radius: 11px;
            border: 1px solid var(--border-strong);
            background: #ffffff;
            color: var(--text-main);
            box-shadow: none;
        }

        .stTextInput input:focus,
        .stTextArea textarea:focus {
            border-color: var(--accent);
            box-shadow: 0 0 0 1px rgba(89, 104, 92, 0.24);
        }

        .stButton button {
            border-radius: 999px;
            border: 1px solid var(--border);
            background: #ffffff;
            color: var(--text-main);
            font-weight: 620;
            min-height: 2.6rem;
            padding-left: 1.05rem;
            padding-right: 1.05rem;
        }

        .stButton button:hover {
            border-color: var(--border-strong);
            color: var(--text-main);
            background: #f8fafc;
        }

        [data-testid="stSidebar"] .stButton button {
            min-height: 2.35rem;
            padding-left: 0.85rem;
            padding-right: 0.85rem;
            width: 100%;
        }

        .stButton button[kind="primary"] {
            background: var(--text-main);
            border-color: var(--text-main);
            color: #ffffff;
        }

        .stButton button[kind="primary"]:hover {
            background: #343431;
            border-color: #343431;
            color: #ffffff;
        }

        div[data-testid="stFileUploader"] section {
            border-radius: 0 0 18px 18px;
            border-color: var(--border);
            border-top: 0;
            background: var(--panel-bg);
            min-height: 92px !important;
            padding: 0.85rem 1rem !important;
            align-items: flex-start;
            box-shadow: var(--shadow);
        }

        div[data-testid="stFileUploader"] section > div {
            min-height: auto !important;
            gap: 0.55rem;
        }

        div[data-testid="stFileUploader"] small {
            margin-top: 0.15rem;
        }

        div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] {
            min-height: 92px !important;
        }

        div[data-testid="stFileUploader"] section:hover {
            border-color: var(--border-strong);
            border-top: 0;
            background: #f8fafc;
        }

        div[data-testid="stDataFrame"] {
            border: 1px solid var(--border);
            border-radius: 14px;
            overflow: hidden;
        }

        .rq-label {
            font-weight: 640;
            color: var(--text-muted);
            padding-top: 0.45rem;
        }

        .rq-control-button {
            margin-top: 0.15rem;
        }

        .section-note {
            color: var(--text-muted);
            font-size: 0.86rem;
            margin-top: -0.35rem;
            margin-bottom: 0.75rem;
        }

        .app-hero {
            margin-bottom: 1.25rem;
            max-width: 860px;
        }

        .app-kicker {
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
            color: var(--text-muted);
            font-size: 0.82rem;
            font-weight: 560;
            margin-bottom: 0.75rem;
            letter-spacing: 0.01em;
        }

        .app-kicker svg,
        .panel-title svg,
        .upload-title svg,
        .metric-card svg {
            width: 17px;
            height: 17px;
            stroke: currentColor;
            stroke-width: 1.75;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
        }

        .workspace-panel {
            border: 1px solid var(--border);
            background: var(--panel-bg);
            border-radius: 18px;
            padding: 1.05rem;
            box-shadow: var(--shadow);
            min-height: 100%;
        }

        .panel-title {
            display: flex;
            align-items: center;
            gap: 0.52rem;
            font-weight: 660;
            color: var(--text-main);
            margin-bottom: 0.25rem;
        }

        .panel-subtitle {
            color: var(--text-muted);
            font-size: 0.9rem;
            line-height: 1.45;
            margin-bottom: 0.95rem;
        }

        .metric-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.62rem;
            margin: 0.85rem 0 0;
        }

        .metric-card {
            border: 1px solid var(--border);
            border-radius: 14px;
            padding: 0.78rem;
            background: #f8fafc;
            min-height: 78px;
        }

        .metric-card svg {
            color: var(--accent);
            margin-bottom: 0.34rem;
        }

        .metric-label {
            color: var(--text-muted);
            font-size: 0.76rem;
            line-height: 1.2;
        }

        .metric-value {
            color: var(--text-main);
            font-weight: 650;
            font-size: 1.02rem;
            margin-top: 0.16rem;
            overflow-wrap: anywhere;
        }

        .upload-shell {
            border: 1px solid var(--border);
            border-bottom: 0;
            background: var(--panel-bg);
            border-radius: 18px 18px 0 0;
            padding: 0.95rem 1rem 0.85rem;
            box-shadow: var(--shadow);
            margin-bottom: -1px;
        }

        .upload-title {
            display: flex;
            align-items: center;
            gap: 0.52rem;
            font-weight: 660;
            color: var(--text-main);
            margin-bottom: 0.22rem;
        }

        .upload-subtitle {
            color: var(--text-muted);
            font-size: 0.9rem;
            line-height: 1.45;
            margin-bottom: 0;
        }

        .action-spacer {
            height: 1.55rem;
        }

        .action-row-note {
            color: var(--text-muted);
            font-size: 0.84rem;
            line-height: 1.45;
            padding-top: 0.55rem;
        }

        .results-spacer {
            margin-top: 1.15rem;
        }

        @media (max-width: 900px) {
            .top-brand {
                left: 3.7rem;
                top: 1.28rem;
                font-size: 1.12rem;
            }

            [data-testid="stSidebar"]::before {
                left: 1.15rem;
                top: 1.15rem;
                font-size: 1.12rem;
            }

            .metric-grid {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }

            .main .block-container {
                padding-left: 1rem;
                padding-right: 1rem;
                padding-top: 1.4rem;
            }

            .upload-shell,
            .workspace-panel {
                padding: 0.9rem;
                border-radius: 15px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def svg_icon(name: str) -> str:
    icons = {
        "spark": '<path d="M12 3l1.8 4.2L18 9l-4.2 1.8L12 15l-1.8-4.2L6 9l4.2-1.8L12 3z"/><path d="M5 14l.9 2.1L8 17l-2.1.9L5 20l-.9-2.1L2 17l2.1-.9L5 14z"/>',
        "file": '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M8 13h8"/><path d="M8 17h5"/>',
        "list": '<path d="M8 6h13"/><path d="M8 12h13"/><path d="M8 18h13"/><path d="M3 6h.01"/><path d="M3 12h.01"/><path d="M3 18h.01"/>',
        "search": '<circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/>',
        "sheet": '<path d="M4 4h16v16H4z"/><path d="M4 10h16"/><path d="M10 4v16"/>',
        "settings": '<path d="M12 15.5A3.5 3.5 0 1 0 12 8a3.5 3.5 0 0 0 0 7.5z"/><path d="M19.4 15a1.7 1.7 0 0 0 .3 1.9l.1.1a2 2 0 1 1-2.8 2.8l-.1-.1a1.7 1.7 0 0 0-1.9-.3 1.7 1.7 0 0 0-1 1.5V21a2 2 0 1 1-4 0v-.1a1.7 1.7 0 0 0-1-1.5 1.7 1.7 0 0 0-1.9.3l-.1.1a2 2 0 1 1-2.8-2.8l.1-.1a1.7 1.7 0 0 0 .3-1.9 1.7 1.7 0 0 0-1.5-1H3a2 2 0 1 1 0-4h.1a1.7 1.7 0 0 0 1.5-1 1.7 1.7 0 0 0-.3-1.9l-.1-.1A2 2 0 1 1 7 4.2l.1.1a1.7 1.7 0 0 0 1.9.3 1.7 1.7 0 0 0 1-1.5V3a2 2 0 1 1 4 0v.1a1.7 1.7 0 0 0 1 1.5 1.7 1.7 0 0 0 1.9-.3l.1-.1A2 2 0 1 1 19.8 7l-.1.1a1.7 1.7 0 0 0-.3 1.9 1.7 1.7 0 0 0 1.5 1h.1a2 2 0 1 1 0 4h-.1a1.7 1.7 0 0 0-1.5 1z"/>',
    }
    return f'<svg viewBox="0 0 24 24" aria-hidden="true">{icons[name]}</svg>'


def render_header() -> None:
    st.markdown(
        f"""
        <div class="top-brand">AQEReview</div>
        <div class="app-hero">
            <div class="app-kicker">{svg_icon("spark")} AQEReview systematic review workspace</div>
            <h1>Citation Screening, Evidence Extraction & Quality Appraisal</h1>
            <div style="color:#64748b; max-width:760px; line-height:1.55; font-size:1rem;">
                Deduplicate RIS/NBIB records, flag clearly irrelevant citations, then extract PDF evidence and run MMAT 2018 quality assessment.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_workspace_panel(
    uploaded_count: int,
    field_count: int,
    question_count: int,
    extraction_result_count: int,
    qa_result_count: int,
    model: str,
) -> None:
    st.markdown(
        f"""
        <div class="workspace-panel">
            <div class="panel-title">{svg_icon("search")} Workflow status</div>
            <div class="panel-subtitle">A compact check of the current batch setup before you run extraction or MMAT quality assessment.</div>
            <div class="metric-grid">
                <div class="metric-card">
                    {svg_icon("file")}
                    <div class="metric-label">PDF files</div>
                    <div class="metric-value">{uploaded_count}</div>
                </div>
                <div class="metric-card">
                    {svg_icon("list")}
                    <div class="metric-label">Structured fields</div>
                    <div class="metric-value">{field_count}</div>
                </div>
                <div class="metric-card">
                    {svg_icon("search")}
                    <div class="metric-label">Research questions</div>
                    <div class="metric-value">{question_count}</div>
                </div>
                <div class="metric-card">
                    {svg_icon("sheet")}
                    <div class="metric-label">Extraction results</div>
                    <div class="metric-value">{extraction_result_count}</div>
                </div>
                <div class="metric-card">
                    {svg_icon("sheet")}
                    <div class="metric-label">MMAT results</div>
                    <div class="metric-value">{qa_result_count}</div>
                </div>
                <div class="metric-card">
                    {svg_icon("settings")}
                    <div class="metric-label">Model</div>
                    <div class="metric-value">{escape(model)}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_upload_intro() -> None:
    st.markdown(
        f"""
        <div class="upload-shell">
            <div class="upload-title">{svg_icon("file")} Source PDFs</div>
            <div class="upload-subtitle">Upload the articles for this extraction run. Files stay local until each PDF is sent to the configured model provider.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_settings() -> tuple[str, str, str]:
    st.subheader("API settings")
    api_key = st.text_input("API key", type="password", help="The app uses this only for the current browser session.")
    base_url = st.text_input("Base URL", value=DEFAULT_BASE_URL)
    model = st.text_input("Model", value=DEFAULT_MODEL)
    return api_key, base_url, model


def restore_default_prompt() -> None:
    st.session_state.prompt_template = DEFAULT_PROMPT_TEMPLATE


def restore_default_mmat_prompt() -> None:
    st.session_state.mmat_prompt_template = DEFAULT_MMAT_PROMPT_TEMPLATE


def restore_default_citation_exclusion_prompt() -> None:
    st.session_state.citation_exclusion_prompt_template = DEFAULT_EXCLUSION_PROMPT_TEMPLATE


def saved_prompt_value(saved_key: str, default_value: str) -> str:
    value = st.session_state.get(saved_key, "")
    if not str(value).strip():
        value = default_value
        st.session_state[saved_key] = value
    return value


def render_prompt_editor_content() -> str:
    saved_value = saved_prompt_value("prompt_template", DEFAULT_PROMPT_TEMPLATE)
    st.markdown(
        '<div class="section-note">Edit the prompt used for article data extraction. The placeholders are filled automatically before each PDF is sent to the model.</div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Restore default extraction prompt",
        key="restore_prompt",
        help="Reset the extraction prompt template to the built-in default.",
        on_click=restore_default_prompt,
    )
    prompt_template = st.text_area(
        "Extraction prompt template",
        value=saved_value,
        height=420,
        help="Keep {structured_fields} and {research_questions} if you want the app to insert the current template fields and RQs at those positions.",
        key="prompt_template_text_area",
    )
    st.session_state.prompt_template = prompt_template
    missing = [
        placeholder
        for placeholder in ("{structured_fields}", "{research_questions}")
        if placeholder not in prompt_template
    ]
    if missing:
        st.info(
            "Missing placeholders will be appended automatically when extraction runs: "
            + ", ".join(missing)
        )
    return prompt_template


def render_prompt_editor() -> str:
    with st.expander("Extraction prompt", expanded=False):
        return render_prompt_editor_content()


def render_citation_exclusion_prompt_editor_content() -> str:
    saved_value = saved_prompt_value(
        "citation_exclusion_prompt_template",
        DEFAULT_EXCLUSION_PROMPT_TEMPLATE,
    )
    st.markdown(
        '<div class="section-note">Edit the prompt used to mark clearly irrelevant citation records. The placeholders are filled automatically before the metadata is sent to the model.</div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Restore default citation exclusion prompt",
        key="restore_citation_exclusion_prompt",
        help="Reset the citation exclusion prompt template to the built-in default.",
        on_click=restore_default_citation_exclusion_prompt,
    )
    prompt_template = st.text_area(
        "Citation exclusion prompt template",
        value=saved_value,
        height=360,
        help="Keep {exclusion_criteria} and {records_json} if you want the app to insert the current criteria and citation records at those positions.",
        key="citation_exclusion_prompt_text_area",
    )
    st.session_state.citation_exclusion_prompt_template = prompt_template
    missing = [
        placeholder
        for placeholder in ("{exclusion_criteria}", "{records_json}")
        if placeholder not in prompt_template
    ]
    if missing:
        st.info(
            "Missing placeholders will be appended automatically when AI exclusion marking runs: "
            + ", ".join(missing)
        )
    return prompt_template


def render_citation_exclusion_prompt_editor() -> str:
    with st.expander("Citation exclusion prompt", expanded=False):
        return render_citation_exclusion_prompt_editor_content()


def render_mmat_prompt_editor_content() -> str:
    saved_value = saved_prompt_value("mmat_prompt_template", DEFAULT_MMAT_PROMPT_TEMPLATE)
    st.markdown(
        '<div class="section-note">Edit the prompt used for MMAT quality assessment. This is separate from the extraction prompt.</div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Restore default MMAT prompt",
        key="restore_mmat_prompt",
        help="Reset the MMAT prompt template to the built-in default.",
        on_click=restore_default_mmat_prompt,
    )
    prompt_template = st.text_area(
        "MMAT prompt template",
        value=saved_value,
        height=420,
        help="Keep {screening_questions} and {mmat_criteria} if you want the app to insert the MMAT 2018 criteria at those positions.",
        key="mmat_prompt_text_area",
    )
    st.session_state.mmat_prompt_template = prompt_template
    missing = [
        placeholder
        for placeholder in ("{screening_questions}", "{mmat_criteria}")
        if placeholder not in prompt_template
    ]
    if missing:
        st.info(
            "Missing placeholders will be appended automatically when quality assessment runs: "
            + ", ".join(missing)
        )
    return prompt_template


def render_mmat_prompt_editor() -> str:
    with st.expander("MMAT assessment prompt", expanded=False):
        return render_mmat_prompt_editor_content()


def render_prompt_settings() -> tuple[str, str, str]:
    citation_exclusion_prompt_template = st.session_state.citation_exclusion_prompt_template
    prompt_template = st.session_state.prompt_template
    mmat_prompt_template = st.session_state.mmat_prompt_template
    with st.expander("Prompt settings", expanded=False):
        st.markdown(
            '<div class="section-note">Choose one AI step to edit. The other prompts stay saved in the background.</div>',
            unsafe_allow_html=True,
        )
        prompt_choice = st.selectbox(
            "Prompt to edit",
            ["Citation exclusion", "Article extraction", "MMAT assessment"],
            key="prompt_editor_choice",
            label_visibility="collapsed",
        )
        if prompt_choice == "Citation exclusion":
            citation_exclusion_prompt_template = render_citation_exclusion_prompt_editor_content()
        elif prompt_choice == "Article extraction":
            prompt_template = render_prompt_editor_content()
        else:
            mmat_prompt_template = render_mmat_prompt_editor_content()
    return citation_exclusion_prompt_template, prompt_template, mmat_prompt_template


def add_research_question() -> None:
    st.session_state.research_questions.append("")


def delete_research_question(index: int) -> None:
    if len(st.session_state.research_questions) == 1:
        st.session_state.research_questions[0] = ""
    else:
        st.session_state.research_questions.pop(index)


def render_research_questions() -> list[str]:
    title_col, add_col = st.columns([0.78, 0.22])
    with title_col:
        st.markdown("**Research questions**")
        st.markdown(
            '<div class="section-note">Use one box for one RQ. The app will label them as RQ1, RQ2, RQ3 in the Excel export.</div>',
            unsafe_allow_html=True,
        )
    with add_col:
        st.markdown('<div class="rq-control-button"></div>', unsafe_allow_html=True)
        st.button("＋", key="add_rq", help="Add a research question", on_click=add_research_question)

    for index in range(len(st.session_state.research_questions)):
        label_col, input_col, delete_col = st.columns([0.16, 0.68, 0.16], vertical_alignment="top")
        with label_col:
            st.markdown(f'<div class="rq-label">RQ{index + 1}</div>', unsafe_allow_html=True)
        with input_col:
            st.session_state.research_questions[index] = st.text_area(
                f"RQ{index + 1}",
                value=st.session_state.research_questions[index],
                key=f"rq_text_{index}",
                height=88,
                label_visibility="collapsed",
                placeholder="Enter one research question",
            )
        with delete_col:
            st.button(
                "×",
                key=f"delete_rq_{index}",
                help=f"Delete RQ{index + 1}",
                on_click=delete_research_question,
                args=(index,),
            )

    return [question.strip() for question in st.session_state.research_questions if question.strip()]


def render_template() -> tuple[list[str], list[str]]:
    st.subheader("Extraction template")
    st.markdown(
        '<div class="section-note">The extraction now uses an exhaustive strategy: it asks the model to include all relevant evidence it can find.</div>',
        unsafe_allow_html=True,
    )
    field_text = st.text_area(
        "Structured fields, one per line",
        value="Study design\nPopulation / sample\nIntervention or exposure\nMain findings\nLimitations",
        height=150,
        key="structured_fields_text",
    )
    questions = render_research_questions()
    return split_lines(field_text), questions


def render_results() -> None:
    if st.session_state.results:
        st.subheader("Extraction results")
        rows = [result_to_flat_row(result) for result in st.session_state.results]
        df = pd.DataFrame(rows)
        st.dataframe(style_results(df), width="stretch")

    if st.session_state.qa_results:
        st.subheader("MMAT quality assessment results")
        rows = [mmat_result_to_summary_row(result) for result in st.session_state.qa_results]
        df = pd.DataFrame(rows)
        st.dataframe(style_results(df), width="stretch")

    if st.session_state.results or st.session_state.qa_results:
        export_bytes = build_excel_export(
            st.session_state.results,
            st.session_state.qa_results,
        )
        st.download_button(
            "Download Excel export",
            data=export_bytes,
            file_name=f"systematic_review_extraction_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if st.session_state.errors:
        st.subheader("Failed PDFs")
        for error in st.session_state.errors:
            st.error(f"{error['file']}: {error['message']}")

    if st.session_state.qa_errors:
        st.subheader("Failed MMAT assessments")
        for error in st.session_state.qa_errors:
            st.error(f"{error['file']}: {error['message']}")


def run_extraction_batch(
    uploaded_files: list[Any],
    api_key: str,
    base_url: str,
    model: str,
    fields: list[str],
    questions: list[str],
    prompt_template: str,
    status: Any,
    progress: Any,
    progress_offset: int = 0,
    progress_total: int | None = None,
) -> None:
    st.session_state.results = []
    st.session_state.errors = []
    total = progress_total or len(uploaded_files)

    for index, uploaded_file in enumerate(uploaded_files, start=1):
        status.info(f"Extracting {uploaded_file.name} ({index}/{len(uploaded_files)})")
        try:
            result = extract_from_pdf(
                uploaded_file=uploaded_file,
                api_key=api_key,
                base_url=base_url,
                model=model,
                fields=fields,
                questions=questions,
                prompt_template=prompt_template,
            )
            st.session_state.results.append(result)
        except Exception as exc:
            st.session_state.errors.append(
                {"file": uploaded_file.name, "message": str(exc)}
            )
        progress.progress((progress_offset + index) / total)


def run_mmat_batch(
    uploaded_files: list[Any],
    api_key: str,
    base_url: str,
    model: str,
    mmat_prompt_template: str,
    status: Any,
    progress: Any,
    progress_offset: int = 0,
    progress_total: int | None = None,
) -> None:
    st.session_state.qa_results = []
    st.session_state.qa_errors = []
    total = progress_total or len(uploaded_files)

    for index, uploaded_file in enumerate(uploaded_files, start=1):
        status.info(f"Assessing MMAT quality for {uploaded_file.name} ({index}/{len(uploaded_files)})")
        try:
            result = assess_quality_from_pdf(
                uploaded_file=uploaded_file,
                api_key=api_key,
                base_url=base_url,
                model=model,
                prompt_template=mmat_prompt_template,
            )
            st.session_state.qa_results.append(result)
        except Exception as exc:
            st.session_state.qa_errors.append(
                {"file": uploaded_file.name, "message": str(exc)}
            )
        progress.progress((progress_offset + index) / total)


def can_run_common(api_key: str, uploaded_files: list[Any] | None) -> bool:
    if not api_key:
        st.warning("Please enter an API key before running.")
        return False
    if not uploaded_files:
        st.warning("Please upload at least one PDF.")
        return False
    return True


def render_citation_metrics() -> None:
    imported_count = int(st.session_state.get("citation_imported_count", 0))
    records = st.session_state.get("citation_records", [])
    duplicate_count = len(st.session_state.get("citation_duplicate_log", []))
    ai_irrelevant_count = sum(1 for record in records if record.get("ai_suggested_exclusion"))

    metric_col_1, metric_col_2, metric_col_3, metric_col_4 = st.columns(4)
    metric_col_1.metric("Imported records", imported_count)
    metric_col_2.metric("Duplicates removed", duplicate_count)
    metric_col_3.metric("Records after deduplication", len(records))
    metric_col_4.metric("AI-marked irrelevant records", ai_irrelevant_count)


def render_citation_screening(
    api_key: str,
    base_url: str,
    model: str,
    exclusion_prompt_template: str,
) -> None:
    st.subheader("Citation screening")
    st.markdown(
        '<div class="section-note">Upload RIS or PubMed NBIB files, remove duplicates, and conservatively mark clearly irrelevant records before full-text extraction.</div>',
        unsafe_allow_html=True,
    )
    citation_files = st.file_uploader(
        "Upload RIS or NBIB citation files",
        type=["ris", "nbib"],
        accept_multiple_files=True,
        key="citation_file_uploader",
        help="You can upload files from multiple databases. PubMed Citation Manager .nbib files are supported.",
    )
    exclusion_text = st.text_area(
        "Exclusion Criteria, one per line",
        key="citation_exclusion_criteria",
        height=120,
        placeholder="Example: Not an empirical study\nExample: Not about the target population",
    )
    exclusion_criteria = split_lines(exclusion_text)

    render_citation_metrics()

    screen_col_1, screen_col_2, screen_col_3, screen_col_4 = st.columns(
        [0.22, 0.27, 0.29, 0.18],
        vertical_alignment="center",
    )
    with screen_col_1:
        run_deduplication = st.button("Deduplicate", help="Import RIS/NBIB records and remove duplicates.")
    with screen_col_2:
        run_ai_marking = st.button("AI mark", help="Mark clearly irrelevant records after deduplication.")
    with screen_col_3:
        run_full_screening = st.button("Deduplicate + AI mark", help="Run deduplication and AI exclusion marking in one step.")
    with screen_col_4:
        if st.button("Clear", help="Clear citation screening results."):
            st.session_state.citation_records = []
            st.session_state.citation_duplicate_log = []
            st.session_state.citation_import_log = []
            st.session_state.citation_errors = []
            st.session_state.citation_ai_prompt_used = ""
            st.session_state.citation_imported_count = 0
            st.session_state.citation_export_timestamp = ""
            st.rerun()

    if run_deduplication:
        if not citation_files:
            st.warning("Please upload at least one RIS or NBIB file.")
        else:
            records, errors, import_log = parse_citation_uploads(citation_files)
            deduplicated_records, duplicate_log = deduplicate_citations(records)
            st.session_state.citation_records = deduplicated_records
            st.session_state.citation_duplicate_log = duplicate_log
            st.session_state.citation_import_log = import_log
            st.session_state.citation_errors = errors
            st.session_state.citation_ai_prompt_used = ""
            st.session_state.citation_imported_count = len(records)
            st.session_state.citation_export_timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            st.success("Deduplication finished.")
            st.rerun()

    if run_ai_marking:
        if not st.session_state.citation_records:
            st.warning("Please run deduplication before AI exclusion marking.")
        elif not exclusion_criteria:
            st.warning("Please enter at least one Exclusion Criterion.")
        elif not api_key:
            st.warning("Please enter an API key before running AI exclusion marking.")
        else:
            status = st.empty()
            progress = st.progress(0)
            try:
                marked_records, prompt_used = mark_citation_exclusions_batched(
                    records=st.session_state.citation_records,
                    exclusion_criteria=exclusion_criteria,
                    api_key=api_key,
                    base_url=base_url,
                    model=model,
                    prompt_template=exclusion_prompt_template,
                    status=status,
                    progress=progress,
                )
                st.session_state.citation_records = marked_records
                st.session_state.citation_ai_prompt_used = prompt_used
                st.session_state.citation_export_timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                status.success("AI exclusion marking finished.")
                st.rerun()
            except Exception as exc:
                status.error(f"AI exclusion marking failed: {exc}")

    if run_full_screening:
        if not citation_files:
            st.warning("Please upload at least one RIS or NBIB file.")
        elif not exclusion_criteria:
            st.warning("Please enter at least one Exclusion Criterion.")
        elif not api_key:
            st.warning("Please enter an API key before running AI exclusion marking.")
        else:
            status = st.empty()
            status.info("Deduplicating citations...")
            records, errors, import_log = parse_citation_uploads(citation_files)
            deduplicated_records, duplicate_log = deduplicate_citations(records)
            st.session_state.citation_imported_count = len(records)
            st.session_state.citation_duplicate_log = duplicate_log
            st.session_state.citation_import_log = import_log
            st.session_state.citation_errors = errors
            try:
                progress = st.progress(0)
                marked_records, prompt_used = mark_citation_exclusions_batched(
                    records=deduplicated_records,
                    exclusion_criteria=exclusion_criteria,
                    api_key=api_key,
                    base_url=base_url,
                    model=model,
                    prompt_template=exclusion_prompt_template,
                    status=status,
                    progress=progress,
                )
                st.session_state.citation_records = marked_records
                st.session_state.citation_ai_prompt_used = prompt_used
                st.session_state.citation_export_timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                status.success("Deduplication and AI exclusion marking finished.")
                st.rerun()
            except Exception as exc:
                st.session_state.citation_records = deduplicated_records
                st.session_state.citation_ai_prompt_used = ""
                st.session_state.citation_export_timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                status.error(f"Deduplication finished, but AI exclusion marking failed: {exc}")

    if st.session_state.citation_records:
        st.markdown("**Screening results**")
        screening_df = pd.DataFrame(
            [citation_to_screening_row(record) for record in st.session_state.citation_records]
        )
        st.dataframe(screening_df, width="stretch")

        timestamp = st.session_state.citation_export_timestamp or datetime.now().strftime("%Y%m%d_%H%M")
        excel_bytes = build_screening_excel_export(
            records=st.session_state.citation_records,
            duplicate_log=st.session_state.citation_duplicate_log,
            import_log=st.session_state.citation_import_log,
            exclusion_criteria=exclusion_criteria,
            ai_prompt_used=st.session_state.citation_ai_prompt_used,
        )
        relevant_records = [
            record
            for record in st.session_state.citation_records
            if not record.get("ai_suggested_exclusion")
        ]
        all_deduplicated_ris = build_ris_export(st.session_state.citation_records)
        relevant_ris = build_ris_export(relevant_records)

        export_col_1, export_col_2, export_col_3 = st.columns(3)
        with export_col_1:
            st.download_button(
                "Download Excel audit file",
                data=excel_bytes,
                file_name=f"citation_screening_audit_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with export_col_2:
            st.download_button(
                "Download RIS: deduplicated + AI relevant only",
                data=relevant_ris,
                file_name=f"deduplicated_ai_relevant_records_{timestamp}.ris",
                mime="application/x-research-info-systems",
            )
        with export_col_3:
            st.download_button(
                "Download RIS: deduplicated only",
                data=all_deduplicated_ris,
                file_name=f"deduplicated_all_records_{timestamp}.ris",
                mime="application/x-research-info-systems",
            )

    if st.session_state.citation_duplicate_log:
        with st.expander("Duplicate log", expanded=False):
            st.dataframe(pd.DataFrame(st.session_state.citation_duplicate_log), width="stretch")

    if st.session_state.citation_import_log:
        with st.expander("Import log", expanded=False):
            st.dataframe(pd.DataFrame(st.session_state.citation_import_log), width="stretch")

    if st.session_state.citation_errors:
        st.subheader("Failed citation imports")
        for error in st.session_state.citation_errors:
            st.error(f"{error['file']}: {error['message']}")


def main() -> None:
    st.set_page_config(page_title="AQEReview", layout="wide")
    apply_custom_style()
    initialise_state()

    with st.sidebar:
        st.markdown('<div class="sidebar-brand">AQEReview</div>', unsafe_allow_html=True)
        api_key, base_url, model = render_settings()
        st.divider()
        citation_exclusion_prompt_template, prompt_template, mmat_prompt_template = render_prompt_settings()
        st.divider()
        fields, questions = render_template()

    render_header()

    render_citation_screening(api_key, base_url, model, citation_exclusion_prompt_template)
    st.divider()

    render_upload_intro()
    uploaded_files = st.file_uploader(
        "Upload PDF articles",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        help="Select multiple PDFs. Browser folder upload is not required in this first version.",
    )

    render_workspace_panel(
        uploaded_count=len(uploaded_files or []),
        field_count=len(fields),
        question_count=len(questions),
        extraction_result_count=len(st.session_state.results),
        qa_result_count=len(st.session_state.qa_results),
        model=model,
    )

    st.markdown('<div class="action-spacer"></div>', unsafe_allow_html=True)

    col_extract, col_mmat, col_full, col_clear = st.columns(
        [0.22, 0.27, 0.22, 0.18],
        vertical_alignment="center",
    )

    with col_extract:
        run_extraction = st.button("Run extraction")
    with col_mmat:
        run_mmat = st.button("Run quality assessment")
    with col_full:
        run_full = st.button("Run full workflow")
    with col_clear:
        if st.button("Clear results"):
            st.session_state.results = []
            st.session_state.errors = []
            st.session_state.qa_results = []
            st.session_state.qa_errors = []
            st.rerun()

    st.markdown(
        '<div class="action-row-note">Results export to one Excel workbook with extraction sheets, MMAT sheets, and the exact prompts used.</div>',
        unsafe_allow_html=True,
    )

    if run_extraction:
        if can_run_common(api_key, uploaded_files):
            if not fields and not questions:
                st.warning("Please enter at least one structured field or research question.")
            else:
                progress = st.progress(0)
                status = st.empty()
                run_extraction_batch(
                    uploaded_files=uploaded_files,
                    api_key=api_key,
                    base_url=base_url,
                    model=model,
                    fields=fields,
                    questions=questions,
                    prompt_template=prompt_template,
                    status=status,
                    progress=progress,
                )
                status.success("Extraction finished.")

    if run_mmat:
        if can_run_common(api_key, uploaded_files):
            progress = st.progress(0)
            status = st.empty()
            run_mmat_batch(
                uploaded_files=uploaded_files,
                api_key=api_key,
                base_url=base_url,
                model=model,
                mmat_prompt_template=mmat_prompt_template,
                status=status,
                progress=progress,
            )
            status.success("MMAT quality assessment finished.")

    if run_full:
        if not can_run_common(api_key, uploaded_files):
            pass
        elif not fields and not questions:
            st.warning("Please enter at least one structured field or research question.")
        else:
            progress = st.progress(0)
            status = st.empty()
            total_steps = len(uploaded_files) * 2
            run_extraction_batch(
                uploaded_files=uploaded_files,
                api_key=api_key,
                base_url=base_url,
                model=model,
                fields=fields,
                questions=questions,
                prompt_template=prompt_template,
                status=status,
                progress=progress,
                progress_offset=0,
                progress_total=total_steps,
            )
            run_mmat_batch(
                uploaded_files=uploaded_files,
                api_key=api_key,
                base_url=base_url,
                model=model,
                mmat_prompt_template=mmat_prompt_template,
                status=status,
                progress=progress,
                progress_offset=len(uploaded_files),
                progress_total=total_steps,
            )
            status.success("Full workflow finished.")

    st.markdown('<div class="results-spacer"></div>', unsafe_allow_html=True)
    render_results()


if __name__ == "__main__":
    main()
