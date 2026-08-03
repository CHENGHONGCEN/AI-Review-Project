import io
import unittest

from openpyxl import load_workbook

import app
import screening


def result(provider: str, decision: str) -> dict:
    return {
        "decision": decision,
        "suggested_action": decision,
        "suggested_inclusion": decision != "exclude",
        "suggested_exclusion": decision == "exclude",
        "matched_criteria": "1",
        "matched_exclusion_criteria": "",
        "exclusion_reason": "",
        "reason": f"{provider} reason",
        "evidence": "evidence",
        "needs_human_review": decision == "unsure",
        "provider": provider,
        "model": f"{provider}-model",
        "completed_at": "2026-07-31T12:00:00",
    }


class ScreeningExportTests(unittest.TestCase):
    def test_comparison_workbook_structure_and_highlighting(self):
        records = [
            {
                "record_id": "C01",
                "title": "Agreement",
                "abstract": "Abstract",
                "screening_results": {
                    "openai": result("openai", "include"),
                    "gemini": result("gemini", "include"),
                },
            },
            {
                "record_id": "C02",
                "title": "Disagreement",
                "abstract": "Abstract",
                "screening_results": {
                    "openai": result("openai", "include"),
                    "gemini": result("gemini", "exclude"),
                },
            },
            {
                "record_id": "C03",
                "title": "Incomplete",
                "abstract": "Abstract",
                "screening_results": {
                    "openai": result("openai", "unsure"),
                },
            },
        ]
        workbook_bytes = app.build_dual_model_comparison_export(
            records=records,
            duplicate_log=[],
            import_log=[],
            inclusion_criteria=["Eligible"],
            exclusion_criteria=["Ineligible"],
            prompts_used={"openai": "same prompt", "gemini": "same prompt"},
            run_metadata={
                "providers": {
                    "openai": {"model": "openai-model", "status": "complete"},
                    "gemini": {"model": "gemini-model", "status": "partial"},
                }
            },
        )
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        self.assertEqual(
            workbook.sheetnames,
            [
                "Comparison Summary",
                "Model Comparison",
                "Duplicate Log",
                "Import Log",
                "Methodology",
            ],
        )
        comparison_sheet = workbook["Model Comparison"]
        headers = {
            cell.value: cell.column
            for cell in comparison_sheet[1]
        }
        self.assertEqual(
            comparison_sheet.cell(3, headers["Comparison"]).value,
            "Disagreement",
        )
        self.assertEqual(
            comparison_sheet.cell(4, headers["Comparison"]).value,
            "Incomplete",
        )
        self.assertEqual(comparison_sheet.cell(3, 1).fill.fgColor.rgb, "00FECACA")
        self.assertEqual(comparison_sheet.cell(4, 1).fill.fgColor.rgb, "00FEF3C7")

        summary = screening.comparison_summary(records)
        self.assertEqual(summary["agreement"], 1)
        self.assertEqual(summary["disagreement"], 1)
        self.assertEqual(summary["incomplete"], 1)

    def test_provider_audits_keep_results_separate(self):
        records = [
            {
                "record_id": "C01",
                "title": "Different decisions",
                "abstract": "Abstract",
                "screening_results": {
                    "openai": result("openai", "include"),
                    "gemini": result("gemini", "exclude"),
                },
            }
        ]
        openai_bytes = app.build_screening_excel_export(
            records,
            [],
            [],
            ["Eligible"],
            ["Ineligible"],
            "prompt",
            provider="openai",
            model="openai-model",
        )
        gemini_bytes = app.build_screening_excel_export(
            records,
            [],
            [],
            ["Eligible"],
            ["Ineligible"],
            "prompt",
            provider="gemini",
            model="gemini-model",
        )
        openai_sheet = load_workbook(io.BytesIO(openai_bytes))["Screening Results"]
        gemini_sheet = load_workbook(io.BytesIO(gemini_bytes))["Screening Results"]
        headers = {
            cell.value: cell.column
            for cell in openai_sheet[1]
        }
        self.assertEqual(openai_sheet.cell(2, headers["AI suggestion"]).value, "Include")
        self.assertEqual(gemini_sheet.cell(2, headers["AI suggestion"]).value, "Exclude")


if __name__ == "__main__":
    unittest.main()
