import io
from datetime import datetime
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def add_rows_to_sheet(
    sheet: Any,
    rows: list[dict[str, str]],
    clean_cell_value: Callable[[Any], Any],
) -> None:
    if not rows:
        return
    headers = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    sheet.append(headers)
    for row in rows:
        sheet.append([clean_cell_value(row.get(header, "")) for header in headers])


def tune_excel_sheet(
    sheet: Any,
    mmat_response_needs_review: Callable[[str], bool],
) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill("solid", fgColor="FCE4E4")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    header_by_column = {
        cell.column: str(cell.value or "").casefold()
        for cell in sheet[1]
    }

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            is_response_cell = "response" in header_by_column.get(cell.column, "")
            if (
                isinstance(cell.value, str)
                and is_response_cell
                and mmat_response_needs_review(cell.value)
            ):
                cell.fill = review_fill

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        longest_line = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines() or [""]:
                longest_line = max(longest_line, len(line))
        sheet.column_dimensions[column_letter].width = min(max(longest_line + 2, 12), 60)

    for row in sheet.iter_rows():
        max_lines = 1
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            wrapped_lines = max(1, len(value) // 55 + 1)
            explicit_lines = value.count("\n") + 1
            max_lines = max(max_lines, wrapped_lines, explicit_lines)
        sheet.row_dimensions[row[0].row].height = min(max(18, max_lines * 15), 120)


def merge_repeated_evidence_cells(sheet: Any) -> None:
    if sheet.max_row < 3:
        return

    merge_columns = [1, 2, 3, 4]
    group_start = 2
    previous_key = None

    for row_index in range(2, sheet.max_row + 2):
        if row_index <= sheet.max_row:
            current_key = (
                sheet.cell(row=row_index, column=1).value,
                sheet.cell(row=row_index, column=2).value,
                sheet.cell(row=row_index, column=3).value,
            )
        else:
            current_key = None

        if previous_key is None:
            previous_key = current_key
            continue

        if current_key != previous_key:
            group_end = row_index - 1
            if group_end > group_start:
                for column_index in merge_columns:
                    sheet.merge_cells(
                        start_row=group_start,
                        start_column=column_index,
                        end_row=group_end,
                        end_column=column_index,
                    )
                    sheet.cell(row=group_start, column=column_index).alignment = Alignment(
                        vertical="center",
                        wrap_text=True,
                    )
            group_start = row_index
            previous_key = current_key


def build_excel_export(
    results: list[dict[str, Any]],
    qa_results: list[dict[str, Any]],
    result_to_flat_row: Callable[[dict[str, Any]], dict[str, str]],
    result_to_evidence_rows: Callable[[dict[str, Any]], list[dict[str, str]]],
    mmat_result_to_summary_row: Callable[[dict[str, Any]], dict[str, str]],
    mmat_result_to_evidence_rows: Callable[[dict[str, Any]], list[dict[str, str]]],
    clean_cell_value: Callable[[Any], Any],
    mmat_response_needs_review: Callable[[str], bool],
    mmat_manual_version: str,
    confidence_needs_review: Callable[[str], bool] | None = None,
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Article Summary"
    summary_sheet["A1"] = "No extraction results"

    summary_rows = [result_to_flat_row(result) for result in results]
    evidence_rows = []
    for result in results:
        evidence_rows.extend(result_to_evidence_rows(result))

    if summary_rows:
        summary_sheet.delete_rows(1, 1)
        add_rows_to_sheet(summary_sheet, summary_rows, clean_cell_value)
        tune_excel_sheet(summary_sheet, mmat_response_needs_review)

    evidence_sheet = workbook.create_sheet("Evidence Excerpts")
    if evidence_rows:
        add_rows_to_sheet(evidence_sheet, evidence_rows, clean_cell_value)
    else:
        evidence_sheet.append(["File names", "title", "research question", "answer_summary", "excerpt", "source_location", "relevance_note"])
    merge_repeated_evidence_cells(evidence_sheet)
    tune_excel_sheet(evidence_sheet, mmat_response_needs_review)

    mmat_summary_sheet = workbook.create_sheet("MMAT Summary")
    mmat_summary_rows = [mmat_result_to_summary_row(result) for result in qa_results]
    if mmat_summary_rows:
        add_rows_to_sheet(mmat_summary_sheet, mmat_summary_rows, clean_cell_value)
    else:
        mmat_summary_sheet.append(
            [
                "File names",
                "title",
                "suitable_for_mmat",
                "study_design_category",
                "S1 response",
                "S2 response",
            ]
        )
    tune_excel_sheet(mmat_summary_sheet, mmat_response_needs_review)

    mmat_evidence_sheet = workbook.create_sheet("MMAT Evidence")
    mmat_evidence_rows = []
    for result in qa_results:
        mmat_evidence_rows.extend(mmat_result_to_evidence_rows(result))
    if mmat_evidence_rows:
        add_rows_to_sheet(mmat_evidence_sheet, mmat_evidence_rows, clean_cell_value)
    else:
        mmat_evidence_sheet.append(
            [
                "File names",
                "title",
                "study_design_category",
                "section",
                "criterion_id",
                "criterion",
                "response",
                "justification",
                "source_location",
            ]
        )
    tune_excel_sheet(mmat_evidence_sheet, mmat_response_needs_review)

    methodology_sheet = workbook.create_sheet("Methodology Prompt")
    methodology_sheet.append(["item", "value"])
    methodology_sheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    methodology_sheet.append(["Extraction prompt used", results[0].get("prompt_used", "not recorded") if results else "not recorded"])
    methodology_sheet.append(["MMAT manual version", mmat_manual_version])
    methodology_sheet.append(["MMAT editable prompt used", qa_results[0].get("mmat_user_prompt_used", "not recorded") if qa_results else "not recorded"])
    methodology_sheet.append(["MMAT full prompt used", qa_results[0].get("mmat_prompt_used", "not recorded") if qa_results else "not recorded"])
    methodology_sheet.append(["Prompt note", "These are the actual prompt texts sent to the AI model for extraction and MMAT quality assessment."])
    tune_excel_sheet(methodology_sheet, mmat_response_needs_review)
    methodology_sheet.column_dimensions["A"].width = 22
    methodology_sheet.column_dimensions["B"].width = 100
    methodology_sheet.row_dimensions[3].height = 240
    methodology_sheet.row_dimensions[5].height = 180
    methodology_sheet.row_dimensions[6].height = 240

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
